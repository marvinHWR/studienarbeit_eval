"""Clean, formatted .xlsx export of a scored run — the human-readable deliverable.

Carries every score/metric the harness computes plus the complete (untruncated)
question and answer text. The Parquet written alongside it remains the full raw
source of truth (contexts, retrieved indices, full corpus, etc. are not duplicated here).
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Excel's hard per-cell character limit; question/gold_answer/answer are plain QA
# text (not the large contexts/full_corpus fields), so this should never trigger.
_EXCEL_MAX_CELL_LEN = 32767

# PubMedQA-only experiment: em/f1 (structural ~0 against the free-text conclusion) and the
# all-zero reasoning_tokens/text_tokens columns are omitted — they carry no signal for a
# yes/no/maybe label task. They remain in the Parquet (raw source of truth); this is only the
# curated human-readable view. answer_correctness/faithfulness_retrieved are conditionally present.
REPORT_COLS = [
    # Reading block first: question -> gold label -> gold long answer -> the model's answer.
    "question", "gold_label", "gold_answer", "answer",
    # Numeric metrics ordered by importance — quality outcomes before cost/latency.
    "pred_label", "label_acc",
    "faithfulness", "faithfulness_retrieved", "answer_relevancy", "answer_correctness",
    "ctx_precision", "ctx_recall",
    "prompt_tokens", "completion_tokens", "ttft_s", "total_latency_s",
    # Identifiers/bookkeeping last (out of the way of the metrics scan). `arm` and `truncated` are
    # omitted — each sheet already IS one arm, and `truncated` is constantly False here.
    "id", "sample", "kb_size",
]

# Columns holding long free text: wide + wrapped. Everything else: compact.
_WIDE_COLS = {"question", "gold_answer", "answer"}
_WIDE_WIDTH = 60
_NARROW_WIDTH = 14

_NUMERIC_FORMAT_COLS = {
    "ctx_precision", "ctx_recall",
    "answer_correctness", "answer_relevancy", "faithfulness", "faithfulness_retrieved",
    "label_acc", "ttft_s", "total_latency_s",
}
_NUMBER_FORMAT = "0.0000"

# Green fill for rows whose predicted label matched gold (a correct answer) — makes the hit/miss
# pattern scannable at a glance. Grey fill + bold for the per-arm aggregate (mean) row.
_CORRECT_FILL = PatternFill("solid", fgColor="C6EFCE")
_AGG_FILL = PatternFill("solid", fgColor="D9D9D9")

# Identifier columns are never averaged into the aggregate row (a mean id/sample is meaningless).
_NON_AGG_COLS = {"id", "sample", "kb_size"}

# One sheet per arm, in this order (unknown arms appended after). Isolating each condition on its
# own tab makes the three-way No-Context / Full-Context / RAG comparison easy to scan.
_ARM_ORDER = ["no_context", "full_context", "rag"]


def _correct_mask(report: pd.DataFrame) -> pd.Series:
    """Boolean per row: did the model get it right? Uses `label_acc` when present (label datasets
    — encodes normalized pred==gold with abstentions handled), else exact match `em`, else all
    False. Index-aligned to `report`."""
    if "label_acc" in report.columns:
        return report["label_acc"].fillna(0.0).astype(float).eq(1.0)
    if "em" in report.columns:
        return report["em"].fillna(0.0).astype(float).eq(1.0)
    return pd.Series(False, index=report.index)


def _sheet_order(arms: list[str]) -> list[str]:
    """Arms in `_ARM_ORDER` first (those present), then any others in first-seen order."""
    present = set(arms)
    ordered = [a for a in _ARM_ORDER if a in present]
    ordered += [a for a in arms if a not in _ARM_ORDER and a not in ordered]
    return ordered


def _write_sheet(writer: pd.ExcelWriter, sheet_name: str, report: pd.DataFrame, cols: list[str]) -> None:
    """Write one formatted sheet (header bold, wide/wrapped text cols, numeric format, freeze +
    auto-filter). Shared by every arm so the layout is identical and defined in one place."""
    report = report.reset_index(drop=True)
    # Display samples 1-based (0-based in the raw Parquet); a human-friendly view only.
    if "sample" in report.columns:
        report["sample"] = report["sample"] + 1
    report.to_excel(writer, index=False, sheet_name=sheet_name)
    ws = writer.sheets[sheet_name]
    last_data_row = ws.max_row  # header is row 1, so data is rows 2..last_data_row

    header_font = Font(bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    for col_idx, col_name in enumerate(cols, start=1):
        letter = get_column_letter(col_idx)
        ws[f"{letter}1"].font = header_font
        ws.column_dimensions[letter].width = _WIDE_WIDTH if col_name in _WIDE_COLS else _NARROW_WIDTH
        if col_name in _WIDE_COLS:
            for row_idx in range(2, last_data_row + 1):
                ws[f"{letter}{row_idx}"].alignment = wrap_alignment
        if col_name in _NUMERIC_FORMAT_COLS:
            for row_idx in range(2, last_data_row + 1):
                ws[f"{letter}{row_idx}"].number_format = _NUMBER_FORMAT

    # Freeze + filter the data block *before* appending the aggregate, so the mean row stays out
    # of the auto-filter range and isn't reordered when the user sorts.
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{last_data_row}"

    # Highlight every correct-answer row (pred label == gold) in green.
    correct = _correct_mask(report)
    for pos, is_correct in enumerate(correct.tolist()):
        if is_correct:
            for col_idx in range(1, len(cols) + 1):
                ws.cell(row=pos + 2, column=col_idx).fill = _CORRECT_FILL

    _append_aggregate_row(ws, report, cols, last_data_row + 1)


def _append_aggregate_row(ws, report: pd.DataFrame, cols: list[str], row_idx: int) -> None:
    """Append one bold, grey-filled per-arm summary row: the column mean of every numeric metric
    (accuracy, scores, tokens, latency), with a labelled first cell. Identifier columns are left
    blank. Gives each arm's headline numbers without scrolling the raw rows."""
    bold = Font(bold=True)
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = _AGG_FILL
        cell.font = bold
        if col_idx == 1:
            cell.value = f"AGGREGATE (mean, n={len(report)})"
            continue
        if col_name in _NON_AGG_COLS or col_name not in report.columns:
            continue
        series = pd.to_numeric(report[col_name], errors="coerce")
        if series.notna().any():
            cell.value = float(series.mean())
            cell.number_format = _NUMBER_FORMAT


# The three yes/no/maybe labels, in the order they appear as summary rows.
_LABELS = ("yes", "no", "maybe")


def _write_summary_sheet(writer: pd.ExcelWriter, df: pd.DataFrame, arms: list[str]) -> None:
    """Cross-arm comparison sheet (arms as columns): headline metric means side by side, plus the
    count of *correct* yes/no/maybe predictions per arm and each class's gold support. Pooled over
    all samples and kb_size sweep points, mirroring the per-sheet AGGREGATE row. Inserted as the
    first tab so the three-way comparison is the workbook's landing page."""
    def _mean(sub: pd.DataFrame, col: str):
        if col not in sub.columns:
            return None
        s = pd.to_numeric(sub[col], errors="coerce")
        return float(s.mean()) if s.notna().any() else None

    def _gold(sub: pd.DataFrame) -> pd.Series:
        return sub["gold_label"].astype(str).str.strip().str.lower()

    def _acc(sub: pd.DataFrame) -> pd.Series:
        return pd.to_numeric(sub["label_acc"], errors="coerce")

    def _correct_over_support(sub: pd.DataFrame, cls: str) -> str:
        """`"<# correct> / <# gold==cls>"` — correct predictions of that class over its support."""
        gold = _gold(sub)
        correct = int(((gold == cls) & (_acc(sub) == 1.0)).sum())
        return f"{correct} / {int((gold == cls).sum())}"

    # (row label, kind, value-fn). kind "int" -> whole number, "float" -> 4-decimal metric,
    # "str" -> written verbatim (used for the "correct / support" cells).
    specs: list[tuple[str, str, object]] = [
        ("n (rows)", "int", lambda s: int(len(s))),
        ("accuracy (label_acc)", "float", lambda s: _mean(s, "label_acc")),
        ("correct total", "int", lambda s: int((_acc(s) == 1.0).sum())),
        ("correct yes", "str", lambda s: _correct_over_support(s, "yes")),
        ("correct no", "str", lambda s: _correct_over_support(s, "no")),
        ("correct maybe", "str", lambda s: _correct_over_support(s, "maybe")),
        ("faithfulness", "float", lambda s: _mean(s, "faithfulness")),
        ("faithfulness_retrieved", "float", lambda s: _mean(s, "faithfulness_retrieved")),
        ("answer_relevancy", "float", lambda s: _mean(s, "answer_relevancy")),
        ("ctx_precision", "float", lambda s: _mean(s, "ctx_precision")),
        ("ctx_recall", "float", lambda s: _mean(s, "ctx_recall")),
        ("prompt_tokens (mean)", "float", lambda s: _mean(s, "prompt_tokens")),
        ("completion_tokens (mean)", "float", lambda s: _mean(s, "completion_tokens")),
        ("ttft_s (mean)", "float", lambda s: _mean(s, "ttft_s")),
        ("total_latency_s (mean)", "float", lambda s: _mean(s, "total_latency_s")),
    ]

    subs = {arm: df[df["arm"] == arm] for arm in arms}
    ws = writer.book.create_sheet("summary", 0)  # index 0 -> first tab

    bold = Font(bold=True)
    ws.cell(row=1, column=1, value="metric").font = bold
    for j, arm in enumerate(arms, start=2):
        c = ws.cell(row=1, column=j, value=str(arm))
        c.font = bold
        c.alignment = Alignment(horizontal="right")

    row_idx = 2
    for label, kind, fn in specs:
        values = [fn(subs[arm]) for arm in arms]
        if all(v is None for v in values):
            continue  # drop rows with no data in any arm (e.g. faithfulness_retrieved when unscored)
        ws.cell(row=row_idx, column=1, value=label).font = bold
        for j, v in enumerate(values, start=2):
            cell = ws.cell(row=row_idx, column=j, value=v)
            if kind == "float" and v is not None:
                cell.number_format = _NUMBER_FORMAT
        row_idx += 1

    ws.column_dimensions["A"].width = 26
    for j in range(2, len(arms) + 2):
        ws.column_dimensions[get_column_letter(j)].width = _NARROW_WIDTH
    ws.freeze_panes = "B2"


def write_excel_report(df: pd.DataFrame, path: Path) -> None:
    """Write the scored frame as one workbook with a sheet per arm (No-Context / Full-Context /
    RAG), each carrying the `REPORT_COLS` subset. If the frame has no `arm` column, falls back to
    a single `results` sheet."""
    cols = [c for c in REPORT_COLS if c in df.columns]

    for col in _WIDE_COLS & set(cols):
        offender = df[col].map(lambda v: isinstance(v, str) and len(v) > _EXCEL_MAX_CELL_LEN)
        if offender.any():
            print(f"WARNING: {offender.sum()} row(s) in column '{col}' exceed Excel's "
                  f"{_EXCEL_MAX_CELL_LEN}-char cell limit and were NOT truncated; "
                  f"Excel may reject or clip these cells on open.")

    _sort_cols = [c for c in ("id", "sample", "kb_size") if c in df.columns]
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        if "arm" not in df.columns:
            report = df[cols]
            if _sort_cols:
                report = report.sort_values(_sort_cols)
            _write_sheet(writer, "results", report, cols)
            print(f"Excel export -> {path}  ({len(report)} rows, 1 sheet)")
            return

        arms = _sheet_order(list(df["arm"].unique()))
        for arm in arms:
            report = df[df["arm"] == arm][cols]
            if _sort_cols:
                report = report.sort_values(_sort_cols)
            _write_sheet(writer, str(arm), report, cols)
        if "gold_label" in df.columns and "label_acc" in df.columns:
            _write_summary_sheet(writer, df, arms)

    print(f"Excel export -> {path}  ({len(df)} rows, "
          f"{len(arms)} arm sheets + summary: {', '.join(arms)})")
