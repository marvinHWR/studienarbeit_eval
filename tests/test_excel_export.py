"""Offline test for the .xlsx report writer: no API key, model download, or network."""
from __future__ import annotations

import openpyxl
import pandas as pd

from sae.reporting.excel_export import write_excel_report


def _header(ws) -> list[str]:
    return [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]


def test_write_excel_report_per_arm_sheets(tmp_path):
    df = pd.DataFrame({
        "id": ["q1", "q2", "q3"],
        "arm": ["rag", "no_context", "full_context"],
        "question": ["Does aspirin reduce risk?", "Is the marker specific?", "Does dosing matter?"],
        "gold_label": ["yes", "no", "maybe"],
        "gold_answer": ["It does.", "It is not.", "Unclear."],
        "answer": ["Yes. It does.", "No. It is not.", "Maybe. Unclear."],
        "pred_label": ["yes", "no", "maybe"],
        "label_acc": [1.0, 1.0, 1.0],
    })

    out = tmp_path / "answers_test.xlsx"
    write_excel_report(df, out)

    assert out.exists()
    wb = openpyxl.load_workbook(out)
    # Cross-arm summary first, then one sheet per arm in fixed order (no_context, full_context,
    # rag); no interleaved "results" sheet.
    assert wb.sheetnames == ["summary", "no_context", "full_context", "rag"]

    # Columns follow the REPORT_COLS allowlist order, filtered to what the frame carries:
    # reading block first, identifiers last. `arm` is dropped — each sheet already IS one arm.
    expected_cols = ["question", "gold_label", "gold_answer", "answer",
                     "pred_label", "label_acc", "id"]
    for name in ("no_context", "full_context", "rag"):
        ws = wb[name]
        assert _header(ws) == expected_cols
        assert ws.freeze_panes == "A2"
        assert ws["A1"].font.bold is True

    # Each row lands on its arm's sheet.
    id_col = expected_cols.index("id") + 1
    assert wb["rag"].cell(row=2, column=id_col).value == "q1"
    assert wb["no_context"].cell(row=2, column=id_col).value == "q2"
    assert wb["full_context"].cell(row=2, column=id_col).value == "q3"
    # ... carrying its own metrics.
    assert wb["rag"].cell(row=2, column=expected_cols.index("label_acc") + 1).value == 1.0


def test_write_excel_report_single_sheet_when_no_arm(tmp_path):
    """Frames without an `arm` column fall back to a single `results` sheet."""
    df = pd.DataFrame({
        "id": ["q1", "q2"],
        "question": ["Does aspirin reduce risk?", "Is the marker specific?"],
        "gold_label": ["yes", "no"],
        "gold_answer": ["It does.", "It is not."],
        "answer": ["Yes. It does.", "No. It is not."],
        "label_acc": [1.0, 0.0],
    })

    out = tmp_path / "answers_noarm.xlsx"
    write_excel_report(df, out)

    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["results"]
    ws = wb["results"]
    header = _header(ws)
    assert ws.cell(row=2, column=header.index("id") + 1).value == "q1"
    assert ws.freeze_panes == "A2"
