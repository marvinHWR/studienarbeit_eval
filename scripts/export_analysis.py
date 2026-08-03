"""Excel deliverable for the analysis results — the presentable companion to `analyze.py`.

`analyze.py` prints its tables and drops raw CSVs into `results/figures/`; this turns those into
one formatted workbook (`results/analysis_<name>.xlsx`) that can be handed over or pasted into the
thesis without reformatting. It computes no statistics of its own — every number is read back from
the CSVs `analyze.py` wrote, plus the descriptive means from the run Parquet, so `analyze.py`
stays the single source of truth. Run it after `analyze.py`; it makes no API calls.

Sheets: overview (setup + headline findings) / mean_by_arm / label_metrics (incl. majority
baseline) / paired_stats (Wilcoxon + Holm) / efficiency.

Usage:  python scripts/export_analysis.py --run results/run_pubmedqa.parquet
"""
from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from sae.metrics.accuracy import macro_f1

# Descriptive means shown on `mean_by_arm`, in reporting order (quality -> retrieval -> cost).
# Filtered to what the run actually carries, so a run without the strict-faithfulness pass or
# without RAGAS simply shows fewer rows.
MEAN_COLS = [
    "label_acc", "faithfulness", "faithfulness_retrieved", "answer_relevancy",
    "answer_correctness", "ctx_precision", "ctx_recall",
    "prompt_tokens", "completion_tokens", "ttft_s", "total_latency_s",
]
# Column order for the paired-stats sheet: what was compared -> how big -> how certain.
STATS_COLS = ["metric", "arm_a", "arm_b", "n", "median_a", "median_b", "effect_size",
              "ci_low", "ci_high", "statistic", "p_value", "p_holm"]

_P_COLS = {"p_value", "p_holm"}
_INT_COLS = {"n", "n_questions", "n_rows"}
_NUMBER_FORMAT = "0.0000"
_P_FORMAT = "0.00E+00"
_HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
_SIG_FILL = PatternFill("solid", fgColor="C6EFCE")     # Holm-significant comparison
_NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")    # baseline / annotation row
_BOLD = Font(bold=True)


def _write_table(writer: pd.ExcelWriter, sheet: str, df: pd.DataFrame,
                 wide_first_col: int = 22) -> None:
    """Write `df` as a formatted sheet: bold filled header, per-type number formats, frozen
    header row. Shared by every table sheet so the workbook looks like one document."""
    df.to_excel(writer, sheet_name=sheet, index=False)
    ws = writer.sheets[sheet]
    for j, col in enumerate(df.columns, start=1):
        letter = get_column_letter(j)
        cell = ws[f"{letter}1"]
        cell.font = _BOLD
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[letter].width = wide_first_col if j == 1 else 15
        if col in _INT_COLS:
            continue
        fmt = _P_FORMAT if col in _P_COLS else _NUMBER_FORMAT
        for row in range(2, ws.max_row + 1):
            if isinstance(ws.cell(row=row, column=j).value, float):
                ws.cell(row=row, column=j).number_format = fmt
    ws.freeze_panes = "A2"


def _overview_rows(df: pd.DataFrame, run: Path, gold: pd.Series,
                   label_tbl: pd.DataFrame, stats: pd.DataFrame) -> list[tuple[str, object]]:
    """Key/value rows for the landing sheet: setup, gold distribution, trivial baseline, and the
    headline findings phrased the way they should be phrased in the text."""
    base_acc = float(gold.value_counts().max() / len(gold))
    base_f1 = macro_f1(["yes"] * len(gold), gold.tolist())
    by_arm = label_tbl.set_index("arm")
    fc, rg, nc = by_arm.loc["full_context"], by_arm.loc["rag"], by_arm.loc["no_context"]
    ptok = df.groupby("arm")["prompt_tokens"].mean()
    acc_fc_rag = stats[(stats["metric"] == "label_acc") & (stats["arm_a"] == "full_context")
                       & (stats["arm_b"] == "rag")].iloc[0]

    return [
        ("SETUP", None),
        ("Run-Datei", str(run)),
        ("Zeilen", int(len(df))),
        ("Fragen", int(df["id"].nunique())),
        ("Samples pro Frage", int(df["sample"].nunique())),
        ("Arme", ", ".join(sorted(df["arm"].unique()))),
        ("kb_size", ", ".join(str(k) for k in sorted(df["kb_size"].unique()))),
        ("Generator (config/default.yaml)", "gemini/gemini-3.5-flash-lite, temp 0.2, seed 42"),
        ("Judge (config/default.yaml)", "anthropic/claude-sonnet-5, temp 0.0"),
        ("Retrieval", "top-k = 4, chunk_level = paragraph"),
        (None, None),
        ("GOLD-VERTEILUNG & TRIVIALE BASELINE", None),
        ("gold yes / no / maybe", " / ".join(str(int(gold.value_counts().get(c, 0)))
                                             for c in ("yes", "no", "maybe"))),
        ("Majority-Baseline Accuracy", round(base_acc, 4)),
        ("Majority-Baseline Macro-F1", round(base_f1, 4)),
        (None, None),
        ("KERNAUSSAGEN", None),
        ("Accuracy RAG vs. Full-Context",
         f"{rg['accuracy']:.3f} vs. {fc['accuracy']:.3f} — kein nachweisbarer Unterschied "
         f"(p = {acc_fc_rag['p_value']:.2f}, CI [{acc_fc_rag['ci_low']:.3f}, "
         f"{acc_fc_rag['ci_high']:.3f}])"),
        ("Macro-F1 RAG vs. Full-Context",
         f"{rg['macro_f1']:.3f} vs. {fc['macro_f1']:.3f} — RAG besser bei den Minderheitsklassen "
         f"(deskriptiv, nicht gepaart testbar)"),
        ("No-Context vs. Baseline",
         f"Accuracy {nc['accuracy']:.3f} liegt UNTER der Majority-Baseline von {base_acc:.3f}"),
        ("Token-Effizienz",
         f"RAG braucht {ptok['rag'] / ptok['full_context'] * 100:.1f} % der Prompt-Tokens von "
         f"Full-Context (Faktor {ptok['full_context'] / ptok['rag']:.1f})"),
        (None, None),
        ("HINWEISE ZUR LESART", None),
        ("faithfulness", "gegen den festen full_corpus gescort, nicht gegen den pro Arm "
                         "tatsaechlich gesehenen Kontext -> ueberschaetzt RAGs Grounding"),
        ("answer_relevancy", "bei kurzen yes/no/maybe-Antworten laengenabhaengig (Confound) -> "
                             "nur mit Vorbehalt berichten"),
        ("ctx_precision / ctx_recall", "deterministisch; die Signifikanzen dort bestaetigen die "
                                       "Konstruktion der Arme, sie sind kein empirischer Befund"),
        ("em / f1", "strukturell ~0 gegen die freitextliche long_answer -> nicht berichtet"),
        ("H3 / H4", "entfallen: kb_size hat nur einen Wert (120), kein Sweep"),
        ("Judge-Validierung", "bewusst nicht durchgefuehrt (Limitation). Daher tragen nur die "
                              "judge-freien Metriken die Aussagen: label_acc, Macro-F1, Tokens"),
        (None, None),
        ("Quelle", "erzeugt von scripts/export_analysis.py aus results/figures/*.csv "
                   "(geschrieben von scripts/analyze.py) + dem Run-Parquet"),
    ]


def _write_overview(writer: pd.ExcelWriter, rows: list[tuple[str, object]]) -> None:
    ws = writer.book.create_sheet("overview", 0)
    for i, (key, value) in enumerate(rows, start=1):
        if key is None:
            continue
        kc = ws.cell(row=i, column=1, value=key)
        kc.font = _BOLD
        kc.alignment = Alignment(vertical="top")
        if value is None:                      # section heading
            kc.fill = _HEADER_FILL
            continue
        vc = ws.cell(row=i, column=2, value=value)
        vc.alignment = Alignment(wrap_text=True, vertical="top")
        if isinstance(value, float):
            vc.number_format = _NUMBER_FORMAT
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 96


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--run", required=True)
    ap.add_argument("--stats-dir", default="results/figures",
                    help="directory analyze.py wrote its CSVs to")
    ap.add_argument("--out", default=None, help="output .xlsx (default: results/analysis_<name>.xlsx)")
    args = ap.parse_args()

    run = Path(args.run)
    sdir = Path(args.stats_dir)
    missing = [f for f in ("paired_stats.csv", "label_metrics.csv", "efficiency.csv")
               if not (sdir / f).exists()]
    if missing:
        ap.error(f"missing {', '.join(missing)} in {sdir} — run scripts/analyze.py --run {run} first")

    out = Path(args.out) if args.out else run.parent / f"analysis_{run.stem.replace('run_', '')}.xlsx"
    df = pd.read_parquet(run)
    stats = pd.read_csv(sdir / "paired_stats.csv")
    label_tbl = pd.read_csv(sdir / "label_metrics.csv")
    eff = pd.read_csv(sdir / "efficiency.csv")
    gold = df.groupby("id")["gold_label"].first().str.strip().str.lower()

    # mean_by_arm: descriptive means, arms as columns (metrics as rows reads better in a report).
    cols = [c for c in MEAN_COLS if c in df.columns]
    means = df.groupby("arm")[cols].mean().T.reset_index().rename(columns={"index": "metric"})

    # label_metrics + the trivial baseline as an explicit comparison row.
    label_out = label_tbl.copy()
    label_out.loc[len(label_out)] = {
        "arm": "MAJORITY BASELINE (immer 'yes')", "n_questions": len(gold),
        "accuracy": float(gold.value_counts().max() / len(gold)),
        "macro_f1": macro_f1(["yes"] * len(gold), gold.tolist()),
        "recall_yes": 1.0, "recall_no": 0.0, "recall_maybe": 0.0, "abstention_rate": 0.0,
    }

    # efficiency + absolute token totals (the cost argument in raw numbers).
    tot = df.groupby("arm")[["prompt_tokens", "completion_tokens"]].sum()
    lat = df.groupby("arm")[["ttft_s", "total_latency_s"]].mean()
    eff_out = eff.set_index("arm").join(tot.rename(columns={
        "prompt_tokens": "prompt_tokens_total", "completion_tokens": "completion_tokens_total",
    })).join(lat).reset_index()
    ref = eff_out.set_index("arm")["prompt_tokens"]["full_context"]
    eff_out["prompt_tokens_vs_full_context"] = eff_out["prompt_tokens"] / ref

    stats_out = stats[[c for c in STATS_COLS if c in stats.columns]].copy()
    stats_out["signifikant_holm_5pct"] = stats_out["p_holm"].lt(0.05).map({True: "ja", False: "nein"})

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        _write_table(writer, "mean_by_arm", means)
        _write_table(writer, "label_metrics", label_out, wide_first_col=32)
        _write_table(writer, "paired_stats", stats_out, wide_first_col=18)
        _write_table(writer, "efficiency", eff_out, wide_first_col=16)
        _write_overview(writer, _overview_rows(df, run, gold, label_tbl, stats))

        # Highlight the Holm-significant comparisons and the appended baseline row.
        ws = writer.sheets["paired_stats"]
        for i, sig in enumerate(stats_out["signifikant_holm_5pct"], start=2):
            if sig == "ja":
                for j in range(1, len(stats_out.columns) + 1):
                    ws.cell(row=i, column=j).fill = _SIG_FILL
        ws = writer.sheets["label_metrics"]
        for j in range(1, len(label_out.columns) + 1):
            ws.cell(row=len(label_out) + 1, column=j).fill = _NOTE_FILL

    print(f"Analysis workbook -> {out}  (overview, mean_by_arm, label_metrics, "
          f"paired_stats [{len(stats_out)} Tests], efficiency)")


if __name__ == "__main__":
    main()
